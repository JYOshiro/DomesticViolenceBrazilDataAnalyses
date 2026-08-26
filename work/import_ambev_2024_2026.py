from datetime import date
from pathlib import Path
import sys

sys.path.insert(0, r"C:\Users\jessi\Documents\Codex\2026-08-24\let\work\duckdb-client-2")
import duckdb
import openpyxl


DB_PATH = Path(r"C:\Users\jessi\iCloudDrive\Master - MBA\2026_T2_DATA6000\Python-analyses\banco_disque100.duckdb")
AMBEV_DIR = Path(r"C:\Users\jessi\iCloudDrive\Master - MBA\2026_T2_DATA6000\Ambev")

SOURCE_PERIODS = [
    (2024, 1), (2024, 2), (2024, 3), (2024, 4),
    (2025, 1), (2025, 2), (2025, 3), (2025, 4),
    (2026, 1), (2026, 2),
]


def extract_period(year: int, quarter: int) -> tuple:
    suffix = str(year)[-2:]
    period_label = f"{quarter}Q{suffix}"
    workbook_path = AMBEV_DIR / f"{period_label} Spreadsheets.xlsx"
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

    brazil = workbook["Brazil"]
    beer_header_row = next(
        row
        for row in range(1, brazil.max_row + 1)
        if brazil.cell(row, 1).value == "R$ million"
        and brazil.cell(row - 1, 1).value == "Brazil Beer"
    )
    beer_column = next(
        column
        for column in range(2, brazil.max_column + 1)
        if str(brazil.cell(beer_header_row, column).value).strip() == period_label
    )

    income_tax = workbook["IR"]
    income_tax_column = next(
        column
        for column in range(2, income_tax.max_column + 1)
        if str(income_tax.cell(1, column).value).strip() == period_label
    )
    effective_tax_rate_row = next(
        row
        for row in range(1, income_tax.max_row + 1)
        if str(income_tax.cell(row, 1).value).strip() == "Effective tax rate"
    )

    beer_volume = brazil.cell(beer_header_row + 1, beer_column).value
    beer_net_sales_per_hl = brazil.cell(beer_header_row + 3, beer_column).value
    effective_corporate_tax_rate = income_tax.cell(effective_tax_rate_row, income_tax_column).value

    if None in (beer_volume, beer_net_sales_per_hl, effective_corporate_tax_rate):
        raise ValueError(f"Missing required values in {workbook_path.name}")

    # The workbook has no Brazil Beer gross-sales or indirect-sales-tax figure.
    return (
        date(year, (quarter - 1) * 3 + 1, 1),
        year,
        quarter,
        f"{year}.0-Q{quarter}",
        float(effective_corporate_tax_rate),
        float(beer_net_sales_per_hl),
        float(beer_volume),
        None,
        f"Ambev {workbook_path.name}; sales-tax percentage unavailable in workbook",
    )


records = [extract_period(year, quarter) for year, quarter in SOURCE_PERIODS]

with duckdb.connect(str(DB_PATH)) as connection:
    for year, quarter in SOURCE_PERIODS:
        connection.execute(
            "DELETE FROM ambev.ambev_quarterly WHERE year = ? AND quarter = ?",
            [year, quarter],
        )

    connection.executemany(
        """
        INSERT INTO ambev.ambev_quarterly (
            period_date,
            year,
            quarter,
            year_quarter,
            effective_corp_tax_rate_pct,
            beer_net_sales_brl_per_hl,
            beer_volume_sold_000_hl,
            brazil_beer_sales_tax_pct_gross_sales,
            official_source
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        records,
    )

print(f"Inserted {len(records)} Ambev quarterly records.")
